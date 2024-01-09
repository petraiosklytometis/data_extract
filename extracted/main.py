import os
import pandas as pd
from openpyxl import load_workbook

def read_excel_file(file_path, sheet_name):
    workbook = load_workbook(filename=file_path, read_only=True)
    df = pd.read_excel(workbook, sheet_name=sheet_name, engine='openpyxl', dtype={"email": str, "telefone_fixo": str, "celular": str})
    return df

def fill_empty_cells(df):
    df["email"].fillna(df["telefone_fixo"], inplace=True)
    df["email"].fillna(df["celular"], inplace=True)
    df["telefone_fixo"].fillna(df["celular"], inplace=True)
    return df

def count_empty_rows(df):
    return df[df[["email", "telefone_fixo", "celular"]].isnull().all(axis=1)].shape[0]

def remove_empty_and_duplicate_rows(df):
    df = df.dropna(how="all", subset=["email", "telefone_fixo", "celular"])
    df = df.drop_duplicates(subset=["email", "telefone_fixo", "celular"])
    return df

def process_files(directory_path):
    files = [file for file in os.listdir(directory_path) if file.endswith(".xlsx")]
    columns_to_extract = ["nome", "email", "telefone_fixo", "celular", "bairro"]

    dfs = []
    empty_rows = []
    total_lines_before = 0

    for file in files:
        sheet_name = load_workbook(filename=os.path.join(directory_path, file), read_only=True).sheetnames[0]
        df = read_excel_file(os.path.join(directory_path, file), sheet_name)

        total_lines_before += len(df)

        df = df[columns_to_extract]
        df = fill_empty_cells(df)

        empty_rows.extend(df[df[["email", "telefone_fixo", "celular"]].isnull().all(axis=1)].to_dict('records'))

        dfs.append(df)

    pd.DataFrame(empty_rows).to_csv(os.path.join(directory_path, "empty_rows.csv"), index=False)

    final_df = pd.concat(dfs)
    final_df = final_df.sort_values(by=["bairro"])

    duplicated_rows = final_df[final_df.duplicated(subset=["email", "telefone_fixo", "celular"], keep='first')]
    duplicated_rows.to_csv(os.path.join(directory_path, "duplicated_rows.csv"), index=False)

    final_df = remove_empty_and_duplicate_rows(final_df)

    final_df.to_csv(os.path.join(directory_path, "data_extracted.csv"), index=False)

    total_final_rows = len(final_df)
    total_duplicated_rows = len(duplicated_rows)
    total_deleted_rows = total_lines_before - total_final_rows

    print(f"Total lines deleted: {total_deleted_rows}")
    print(f"Total empty lines before extraction: {len(empty_rows)}")
    print(f"Total duplicated lines: {total_duplicated_rows}")
    print(f"Total remaining lines after deletion: {total_final_rows}")

def main():
    directory_path = "./dataset/"
    process_files(directory_path)

if __name__ == "__main__":
    main()
